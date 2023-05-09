from django.shortcuts import render, redirect, get_object_or_404, HttpResponseRedirect
from django.http import JsonResponse
from OrgDash.models import Skate, Invitation, Player, PlayerGroup, InviteList, Waitlist, LightTeam, DarkTeam, Bench
from django.contrib import messages
from django.contrib.auth.models import User
from django.views.generic import CreateView, DetailView, DeleteView, UpdateView, ListView
from OrgDash.forms import (CreateEventForm, UpdateEventForm, CreateInviteForm, CreatePlayerForm,
PlayerUpdateForm, InviteUpdateForm, InviteWaitlistForm,EventRepeatForm, InitEventRepeatForm, UploadSheetForm, 
)
# from OrgDash.models import UploadSheet
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from django.db.models import Q
from django.urls import reverse_lazy, reverse
from django.utils import timezone









# Create your views here.

def OrganizerDashboard(request):
    today=timezone.now()
    active_user = request.user.pk
    todays_events = Skate.objects.filter(Q(host= active_user) & Q(date=today))
    upcoming_events = Skate.objects.filter(Q(host= active_user) & Q(date__gt= today))
    past_events = Skate.objects.filter(Q(host= active_user) & Q(date__lt= today))
    context = {'upcoming_events' : upcoming_events, 'past_events': past_events, 'todays_events': todays_events}
    return render (request,'OrgDash/dash_base.html',context)
   

##################SKATES/EVENTS##########################

class SkateCreateView(CreateView):
    template_name = 'OrgDash/Skates/create_event.html'
    success_url = reverse_lazy('OrgDash:organizer_dashboard')
    form_class= CreateEventForm
    model = Skate

    def get_initial(self):
        return {'host': self.request.user}
    
    def get_success_url(self):
            if self.object.recurring_event == True:
                return reverse_lazy('OrgDash:skate_repeat_settings', args=[self.object.pk])
            else:
                return super().get_success_url()
            
    # def dispatch(self, request, *args, **kwargs):
    #     messages.success(request, 'Your form was submitted successfully!')
    #     return super().dispatch(request, *args, **kwargs)


class SkateRepeatUpdate(UpdateView):
    model = Skate
    template_name = 'OrgDash/Skates/skate_repeat_form.html'
    form_class = EventRepeatForm

    def get_success_url(self):
        return reverse('OrgDash:event_detail', args=[self.object.pk])
    
    def get_form_class(self):
        modelform = super().get_form_class()
        modelform.base_fields['group_to_invite'].limit_choices_to = {'created_by': self.request.user}
        return modelform
    

class InitSkateRepeatUpdate(UpdateView):
    model = Skate
    template_name = 'OrgDash/Skates/skate_repeat_form.html'
    form_class = InitEventRepeatForm

    def get_success_url(self):
        return reverse('OrgDash:event_detail', args=[self.object.pk])
    
    def get_form_class(self):
        modelform = super().get_form_class()
        modelform.base_fields['group_to_invite'].limit_choices_to = {'created_by': self.request.user}
        return modelform
    
    def get_initial(self):
        return {'recurring_event': True} 

    
class SkateDeleteView(DeleteView): 
    model = Skate
    template_name = 'OrgDash/confirm_delete.html'
    success_url = reverse_lazy('OrgDash:organizer_dashboard')

    
class EventUpdateView(UpdateView):
    model = Skate
    form_class = UpdateEventForm
    template_name = 'OrgDash/update.html'
    
    def get_success_url(self):
        return reverse('OrgDash:event_detail', args=[self.object.pk])


def TeamsView(request, pk):
    from OrgDash.team_sort import SortTeams, SetTeams, RemoveFromTeam, AddToTeam
    active_user = request.user.pk
    active_event = Skate.objects.get(pk=pk)
    guests = Invitation.objects.filter(Q(host= active_user) & Q(event= active_event) & Q(will_you_attend= 'Yes'))
    skaters = active_event.player_guests.all()
    goalies = active_event.goalie_guests.all()
    
   
    #Take player skill and make teams
    player_data =[((guest), float(guest.skill)) for guest in skaters]
    goalie_data = [((guest), float(guest.skill)) for guest in goalies]
    
    # #set light team
    light_team_tup = LightTeam.objects.get_or_create(event=active_event)
    light_team_obj = light_team_tup[0]
    light_team_obj.get_total_skill()
    
    # set dark team
    dark_team_tup = DarkTeam.objects.get_or_create(event=active_event)
    dark_team_obj = dark_team_tup[0]
    dark_team_obj.get_total_skill()
    
    # set bench
    bench_obj = Bench.objects.get_or_create(event=active_event)
    bench_obj[0].SetBench(guests, light_team_obj, dark_team_obj)
   
    total_skill = (float(dark_team_obj.skill) + float(light_team_obj.skill))
    context = {'light_team': light_team_obj.team.all(), 'dark_team': dark_team_obj.team.all(), 'total_skill':total_skill, 'active_event':active_event,
                'light_team_skill' : light_team_obj.skill, 'dark_team_skill': dark_team_obj.skill, 'bench_members': bench_obj[0].bench_members.all()}
    
    
    if request.method == 'POST':
        sort_action = request.POST.get('auto_sort')
        light_form_action = request.POST.get('light_form_action')
        dark_form_action = request.POST.get('dark_form_action')
        bench_action = request.POST.get('bench_action')
        
        if sort_action == 'auto_sort':
            teams = SortTeams(player_data, goalie_data)
            SetTeams(teams, active_event)
            
          
        if light_form_action == 'remove_player_light':
            selected_player_ids = request.POST.getlist('light_player')
            RemoveFromTeam(light_team_obj, selected_player_ids)
            
        if dark_form_action == 'remove_player_dark':
            selected_player_ids = request.POST.getlist('dark_player')
            RemoveFromTeam(dark_team_obj, selected_player_ids)
            
        if bench_action == 'add_to_light':
            selected_player_ids = request.POST.getlist('bench_player')
            AddToTeam(light_team_obj, selected_player_ids)
            
        elif bench_action == 'add_to_dark':
            selected_player_ids = request.POST.getlist('bench_player')
            AddToTeam(dark_team_obj, selected_player_ids)

        return redirect(reverse('OrgDash:make_teams', args=[active_event.pk]), context)
    else:
    
        return render (request,'OrgDash/Skates/make_teams2.html',context)



def EventDash(request, pk): 
    active_user = request.user.pk
    active_event = Skate.objects.get(pk=pk)
    my_player_groups = PlayerGroup.objects.filter(created_by = active_user).all()
    invite_list_obj, created = InviteList.objects.get_or_create(event= active_event)
    if created == False:
        invite_list_obj.guests.clear()
    
    all_invited = Invitation.objects.filter(Q(host= active_user) & Q(event=active_event))
    
    player_list = active_event.player_guests.all()
    goalie_list = active_event.goalie_guests.all()
    wait_list_obj, created = Waitlist.objects.get_or_create(event=active_event)
    wait_list = wait_list_obj.guests.all()
    spots_left = active_event.max_players - len(active_event.player_guests.all())
    goalie_spots_left = active_event.max_goalies - len(active_event.goalie_guests.all())
    
    
    
    context = {'event': active_event, 'guest_list': player_list, 'goalie_list':goalie_list, 'guest_num':(len(player_list) + len(goalie_list)), 
               'spots_left': spots_left, 'goalie_spots_left': goalie_spots_left, 'player_groups':my_player_groups,
                  'wait_list':wait_list, 'wait_num':len(wait_list), 'all_invited': all_invited }
    if request.method == "POST":
        group_id = request.POST.get('invite_group')
        selected_group_members = PlayerGroup.objects.get(id=group_id).members.all()
        for member in selected_group_members:
            invite_list_obj.guests.add(member)
    
        return redirect(reverse('OrgDash:finalize_invites' ,args=[active_event.pk]), context)
    else:
        return render(request, 'OrgDash/Skates/event_detail2.html', context)




def FinalizeRosters(request, pk):
    current_event = Skate.objects.get(pk=pk)
    dark_team = DarkTeam.objects.get(event=current_event)
    light_team = LightTeam.objects.get(event=current_event)
    dark_team_members = dark_team.team.all()
    light_team_members = light_team.team.all()
    
    player_emails = [player.email for player in light_team_members] +[player.email for player in dark_team_members]
    
   
    event_host = current_event.host
    context= {'light_team_members': light_team_members, 'dark_team_members':dark_team_members, 'event':current_event}
    if request.method == 'POST':
        message_field = request.POST['tinymce_email_invite']
        
        
        
        subject = "Rosters for" + event_host.first_name + "'s event at " + current_event.location
        message = message_field 
        recipient_list = player_emails
        recipient_list.append(current_event.host.email)
        try:
            send_mail(subject, message,  'pickuphockey1@gmail.com', recipient_list, html_message=message_field)

            if len(recipient_list) == 1:
                confirm_message = str(len(recipient_list)) + ' Email Sent'
            else:
                confirm_message = str(len(recipient_list)) + ' Emails Sent'
            messages.success(request, confirm_message )
        except:
            messages.error(request, "Error sending email")
        
        
        return redirect(reverse('OrgDash:event_detail' ,args=[current_event.pk]))
    else:
        return render(request, 'OrgDash/Skates/finalize_rosters.html', context)

##########################INVITATIONS################################


class CreateInvite(CreateView): 
    template_name = 'OrgDash/Invites/create_invite.html'
    form_class= CreateInviteForm
    model = Invitation

    def get_form_class(self):
        modelform = super().get_form_class()
        modelform.base_fields['guest'].limit_choices_to = {'created_by': self.request.user}
        return modelform

    def get_success_url(self):
        return reverse('OrgDash:event_detail', args=[str(self.kwargs['pk'])])

    def get_initial(self):
        return {'host': self.request.user, 'event': self.kwargs['pk']}

def RespondToIinvite(request, pk): 
    current_invite = Invitation.objects.get(pk=pk)
    link = reverse('OrgDash:update_invite', kwargs={'pk': current_invite.pk})

    active_event = current_invite.event
   
    waitlist, created = Waitlist.objects.get_or_create(event=active_event)
    context = {'guest':current_invite.guest, 'event':active_event, 'link':link}

    if (current_invite.guest.goalie and active_event.goalie_full) or (not current_invite.guest.goalie and active_event.player_full):
        form_class = InviteWaitlistForm
    else:
        form_class = InviteUpdateForm
    if request.method == 'POST':
        form = form_class(request.POST, instance=current_invite)
        if form.is_valid():
            current_invite.update_event(waitlist)
            current_invite.check_full()
            waitlist.notify_open_spot(current_invite)
            form.save()
            
            if current_invite.will_you_attend == "Yes":
                html_message = render_to_string('OrgDash/emails/RSVP_yes.html', context)
                subject = "You have responded yes to " + str(active_event.host) + "'s event"
            elif current_invite.will_you_attend == "No":
                html_message = render_to_string('OrgDash/emails/RSVP_no.html', context)
                subject = "You have responded no to " + str(active_event.host) + "'s event"
            elif current_invite.will_you_attend == "Waitlist":
                html_message = render_to_string('OrgDash/emails/add_to_waitlist.html', context)
                subject = "You have been added to the waitlist for " + str(active_event.host) + "'s event"
            send_mail(subject, 'message',active_event.host.email, [current_invite.guest.email] , html_message=html_message)
         
            return redirect('OrgDash:invite_confirm_landing', pk=current_invite.pk)
            

    else:
        form = form_class(instance=current_invite)
        if form_class == InviteWaitlistForm:
            return render(request, 'OrgDash/Invites/waitlist_response.html', {'form': form, 'event':active_event})
    
    return render(request, 'OrgDash/Invites/invite_response.html', {'form': form, 'event':active_event})
    

def InviteConfirm(request, pk):
    current_invite = Invitation.objects.get(pk=pk)
    link = reverse('OrgDash:update_invite', kwargs={'pk': current_invite.pk})
    active_event = current_invite.event
    context = {'event':active_event, 'invite':current_invite, 'link':link}
    return render(request, 'OrgDash/Skates/invite_response_landing.html', context)
        

class DeleteInvite(DeleteView): #TODO should you be allowed to delete invites? might cause issues
    model = Invitation
    template_name = 'OrgDash/confirm_delete.html'

    def get_success_url(self):
        current_event = Invitation.objects.get(pk=self.kwargs['pk'])
        event_pk =current_event.event.pk
        return reverse_lazy('OrgDash:event_detail',args = [event_pk])



def AddToInviteList(request, pk):
    active_user = request.user.pk
    active_event = Skate.objects.get(pk=pk)
    my_players = Player.objects.filter(created_by=active_user)
    # my_groups = PlayerGroup.objects.filter(created_by=active_user)
    existing_invites = Invitation.objects.filter(event=active_event)
    player_already_invited =[player.guest for player in existing_invites]
    
    invite_list_tup = InviteList.objects.get_or_create(event= active_event)
    if invite_list_tup[1] == False:
        invite_list_tup[0].guests.clear()
        # for player in player_already_invited:
        #     invite_list_tup[0].guests.remove(player)
   
    invite_list =  list(invite_list_tup[0].guests.all())
    yet_to_invite = []
    for player in my_players:
        if player not in player_already_invited:
                yet_to_invite.append(player)
    
    
    context={
    'active_event':active_event, 'yet_to_invite':yet_to_invite, 
     'invite_list':invite_list, 'inv_obj': invite_list_tup[0]
    
    }
   
    if request.method == 'POST':
        # form_action = request.POST.get('form_action')
        # if form_action == 'add_player':
        selected_player_ids = request.POST.getlist('selected_players_to_add')
        for player_id in selected_player_ids:
            
            invite_list_tup[0].guests.add(player_id)

        # elif form_action == 'remove_player':
        #     selected_player_ids = request.POST.getlist('selected_players_to_remove')
        #     for player_id in selected_player_ids:
        #         invite_list_tup[0].guests.remove(player_id)

        # else: 
        #     group_id = request.POST.get('add_group')
        #     selected_group_members = PlayerGroup.objects.get(id=group_id).members.all()
        #     for member in selected_group_members:
        #         invite_list_tup[0].guests.add(member)
         
        return redirect(reverse('OrgDash:finalize_invites', args=[active_event.pk]), context)
    else:  
        return render(request, 'OrgDash/Invites/invitedash2.html', context )


def FinalizeInvites(request, pk): 
    active_event = Skate.objects.get(pk=pk)
    invite_list_obj = InviteList.objects.get(event=active_event)
    invite_list = invite_list_obj.guests.all()
    num_of_guests = len(invite_list)
    player_emails = [player.email for player in invite_list]
    current_event = invite_list_obj.event
    if request.method == 'POST': 
        message_field = request.POST['tinymce_email_invite']
        invite_list_obj.create_invites()
        subject = "Invitation to " + str(current_event.host) + "'s event at " + current_event.location
       
        recipient_list = player_emails
        
        
        for guest in invite_list:
            invite = Invitation.objects.get(Q(event=active_event) & Q (guest=guest))
            link = reverse('OrgDash:update_invite', kwargs={'pk': invite.pk})
           
            context = {'message': message_field, 'event': current_event, 'guest':guest, 'link':link}
            html_message = render_to_string('OrgDash/emails/invitation_email.html', context)
            
            send_mail(subject, 'message', active_event.host.email, [guest.email] , html_message=html_message)

        
        
        if len(recipient_list) == 1:
            confirm_message = str(len(recipient_list)) + ' Invitation Sent'
        else:
            confirm_message = str(len(recipient_list)) + ' Invitations Sent'
        messages.success(request, confirm_message )
        
        
        return redirect(reverse('OrgDash:event_detail' ,args=[invite_list_obj.event.pk]))
    else:
        return render(request, 'OrgDash/Invites/finalize_invites.html', {'inv_obj': invite_list_obj, 'invite_list':invite_list, 'num_of_guests': num_of_guests, 'event':active_event})




####################PLAYERS##############################


class CreatePlayer(CreateView): 
    template_name = 'OrgDash/Players/create_player.html'
    form_class= CreatePlayerForm
    model = Player
    success_url = reverse_lazy('OrgDash:organizer_dashboard')

    def get_initial(self):
        return {'created_by': self.request.user}

    def get_success_url(self):
        return reverse('OrgDash:player_dash')


class PlayerUpdateView(UpdateView):
    model = Player
    form_class = PlayerUpdateForm
    template_name = 'OrgDash/Players/update_player.html'
    
    def get_success_url(self):
        return reverse('OrgDash:player_dash')


class PlayerDeleteView(DeleteView): 
    model = Player
    template_name = 'OrgDash/confirm_delete.html'
    success_url = reverse_lazy('OrgDash:player_dash')

class PlayerDetail(DetailView):
    model = Player
    context_object_name = 'player'
    template_name = 'OrgDash/Players/player_detail.html'


class PlayerListiview(ListView):
    model = Player
    template_name = 'OrgDash/Players/player_list.html'

    def get_queryset(self):
        all_players = super().get_queryset()
        my_players = all_players.filter(created_by = self.request.user)
        return my_players
    

def Playergroups(request): 
    active_user = request.user
    my_players = Player.objects.filter(created_by=active_user.pk)
    my_groups = PlayerGroup.objects.filter(created_by=active_user.pk)
    context={'my_players': my_players, 'my_groups': my_groups }
    if request.method == 'POST':
        selected_player_ids = request.POST.getlist('selected_players')
        group_name = request.POST.get('group_name')
        selected_player_objs = []
        for player_id in selected_player_ids:
            player = Player.objects.get(pk=player_id)
            selected_player_objs.append(player)
        group_data = {'created_by': active_user, 'name':group_name}
        group = PlayerGroup(**group_data)
        group.save()
        group.members.set(selected_player_objs)
        return redirect(reverse('OrgDash:player_dash'))
    else:

        return render(request, 'OrgDash/Players/create_player_group2.html', context)
    
class PlayerGroupDetail(DetailView):
    model= PlayerGroup
    context_object_name = 'group'
    template_name = 'OrgDash/Players/group_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['members'] = self.object.members.all()
        return context


class PlayerGroupDelete(DeleteView):
    model = PlayerGroup
    template_name = 'OrgDash/confirm_delete.html'
    success_url = reverse_lazy('OrgDash:player_dash')


def UpdatePlayerGroup(request,pk): 
    active_user = request.user
    my_players = Player.objects.filter(created_by=active_user.pk)
    current_group = PlayerGroup.objects.get(pk=pk)
    members = current_group.members.all()
    member_emails = [member.email for member in members]
    non_members =[]
    for player in my_players:
        if player.email not in member_emails:
            non_members.append(player)

    if request.method == 'POST':
        member_ids = request.POST.getlist('members')
        selected_player_objs =[]
        for id in member_ids:
            member = Player.objects.get(pk=id)
            selected_player_objs.append(member)
        current_group.members.set(selected_player_objs)
        return redirect(reverse('OrgDash:player_group_update', args=[current_group.pk]))

    context={'non_members': non_members, 'members':members, 'current_group': current_group}
    return render(request, 'OrgDash/Players/update_group2.html', context)


def PlayerDash(request):
    my_players = Player.objects.filter(created_by=request.user)
    my_groups = PlayerGroup.objects.filter(created_by=request.user)
    my_goalies = [player for player in my_players if player.goalie == True]

    return  render(request, 'OrgDash/Players/player_dash.html', {'my_groups': my_groups, 'my_players':my_players, 'my_goalies':my_goalies})



def UploadSheet(request):
    import openpyxl
    my_players = Player.objects.filter(created_by = request.user)
    player_emails = [player.email for player in my_players]
    if request.method == 'POST':
        form = UploadSheetForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_sheet = form.cleaned_data['file']
            print(uploaded_sheet)
            
            wb= openpyxl.load_workbook(uploaded_sheet)
            sheet= wb.active

            max_row=sheet.max_row
            max_column=sheet.max_column
            player_count = 0
            for i in range(2,max_row+1): #assumes first row of sheet has labels, data starts at row 2
                player_row_values = []
                for j in range(1,max_column+1):
                    cell_obj=sheet.cell(row=i,column=j)  
                    player_row_values.append(cell_obj.value)
                
                player_keys = ['first_name', 'last_name', 'email', 'skill']
                player_data = dict(zip(player_keys, player_row_values))
                player_data['created_by']=request.user
                    
                player = Player(**player_data)
                player_count += 1
                if player.email in player_emails:
                    player_count -= 1
                    
                    
                else:
                   player.save()
            if player_count != 1:
                message= str(player_count) + ' Players added'
            else:
                message = str(player_count) + ' Player added'
            messages.success(request, message)
        else:
            messages.error(request, "Error while creating players" )
            
            
        return redirect('OrgDash:player_dash')
    else:
        form = UploadSheetForm()
        return render(request, 'OrgDash/Players/upload_sheet.html', {'form': form}) 
    














 
    

            
    








    
    