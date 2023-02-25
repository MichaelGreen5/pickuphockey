
def PlayerDataFromExcel(file, start_row):
    import openpyxl

    wb= openpyxl.load_workbook(file)
    sheet= wb.active


    max_row=sheet.max_row

    max_column=sheet.max_column
    player_data = []
    for i in range(start_row,max_row+1):
        for j in range(1,max_column+1):
            cell_obj=sheet.cell(row=i,column=j)  
            player_data.append(cell_obj.value)
    return player_data









        